import os
import pandas as pd
import numpy as np
from semopy import Model, Optimizer
import pydot
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from IPython.display import Image, display

def sem_analysis(model_desc, hypothesis_criteria, p_value_threshold=0.05, dependent_variable='Yvar_USE_AI_Work',
                 segment_results='N', segmentation_column='Generation', filter_column=None, filter_values=None):
    
    # Paths
    current_directory = os.path.dirname(os.getcwd())
    data_dir = os.path.join(current_directory, '01-data')
    summary_dir = os.path.join(current_directory, '04-summary')
    charts_dir = os.path.join(current_directory, '02-charts')

    # Ensure directories exist
    os.makedirs(summary_dir, exist_ok=True)
    os.makedirs(charts_dir, exist_ok=True)

    # Load dataset
    excel_path = os.path.join(data_dir, 'TAM_DEF.xlsx')
    try:
        df = pd.read_excel(excel_path)
        print("Dataset loaded successfully.")
    except FileNotFoundError:
        print(f"Dataset not found at {excel_path}")
        return

    # Apply filtering
    if filter_column and filter_values is not None:
        if not isinstance(filter_values, (list, tuple, np.ndarray)):
            filter_values = [filter_values]
        df = df[~df[filter_column].isin(filter_values)]
        print(f"Filtered out {filter_values} from {filter_column} column.")

    # Segmentation setup
    if segment_results.upper() == 'Y':
        segments = df[segmentation_column].unique()
        print(f"Segments identified: {segments}")
    else:
        segments = ['Entire Dataset']
        df['Entire Dataset'] = 'Entire Dataset'  # Dummy column for uniform processing

    # Initialize results storage
    all_results = []

    # Processing each segment
    for segment in segments:
        print(f"\nProcessing segment: {segment}")
        segment_df = df[df[segmentation_column] == segment] if segment_results.upper() == 'Y' else df

        # Building and fitting the SEM model
        try:
            model = Model(model_desc)
            model.load_dataset(segment_df)
            optimizer = Optimizer(model)
            optimizer.optimize()
            print(f"SEM model optimized for segment: {segment}")
        except Exception as e:
            print(f"Error in model optimization for segment {segment}: {e}")
            continue

        # Extracting results
        estimates = model.inspect()
        estimates = estimates[['lval', 'op', 'rval', 'Estimate', 'Std. Err', 'z-value', 'p-value']]
        
        # Convert z-value and p-value columns to numeric, forcing errors to NaN
        estimates['z-value'] = pd.to_numeric(estimates['z-value'], errors='coerce')
        estimates['p-value'] = pd.to_numeric(estimates['p-value'], errors='coerce')
        
        # Mark significant results based on z-values
        estimates['Significant'] = (estimates['z-value'].abs() > 1.96)
        estimates['Segment'] = segment
        all_results.append(estimates)

        # Generating visualization using pydot
        try:
            graph = pydot.Dot(graph_type='digraph', rankdir='LR')  # 'LR' for left to right layout
            graph.set_size('"16,24!"')  # Increase size for better readability

            # Add nodes with larger font size for readability
            nodes = set(estimates['lval']).union(set(estimates['rval']))
            for node in nodes:
                color = 'lightblue' if node != dependent_variable else 'lightgreen'
                graph.add_node(pydot.Node(node, style='filled', fillcolor=color, fontsize='24', fontname='Arial'))

            # Add edges with color-coded lines and adjusted thickness
            for _, row in estimates.iterrows():
                source = row['rval']
                target = row['lval']
                label = f"{row['Estimate']:.2f}\n(z={row['z-value']:.2f}, p={row['p-value']:.4f})"
                if row['Significant']:
                    color = 'green'
                    penwidth = '3'  # Thicker line for significant paths
                else:
                    color = 'red'
                    penwidth = '1'  # Thinner line for non-significant paths
                fontcolor = 'black'  # Use black text for better readability
                edge = pydot.Edge(source, target, label=label, color=color, fontsize='20', fontcolor=fontcolor, style='solid', penwidth=penwidth, fontname='Arial')
                graph.add_edge(edge)

            # Save and display the graph
            output_file = os.path.join(charts_dir, f'SEM_Visualization_{segment}.png')
            graph.write_png(output_file)
            display(Image(output_file))
            print(f"Visualization saved for segment {segment} at {output_file}")
        except Exception as e:
            print(f"Error in visualization for segment {segment}: {e}")
            continue

    # Combining results
    results_df = pd.concat(all_results, ignore_index=True)

    # Create the Excel file with hypothesis summary
    output_excel = os.path.join(summary_dir, 'SEM_Results.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "Hypotheses Summary"

    # Add headers
    headers = ['Segment', 'Hypothesis', 'p-value', 'Estimate', 'Std. Err', 'z-value', 'Result']
    ws.append(headers)

    # Define formatting
    accepted_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    rejected_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Populate the rows with results
    for segment in segments:
        segment_estimates = results_df[results_df['Segment'] == segment]
        for hyp, path in hypothesis_criteria:
            match = segment_estimates[
                (segment_estimates['lval'] == path.split(' ~ ')[0]) &
                (segment_estimates['rval'] == path.split(' ~ ')[1])
            ]
            if not match.empty:
                row = match.iloc[0]
                result = 'Accepted' if row['Significant'] else 'Rejected'
                fill = accepted_fill if result == 'Accepted' else rejected_fill
                data_row = [segment, hyp, row['p-value'], row['Estimate'], row['Std. Err'], row['z-value'], result]
                ws.append(data_row)
                
                # Apply formatting to the last row added
                for cell in ws[ws.max_row]:
                    cell.fill = fill
                    cell.font = bold_font if result == 'Accepted' else Font()
                    cell.alignment = center_alignment
            else:
                data_row = [segment, hyp, 'N/A', 'N/A', 'N/A', 'N/A', 'Path Not Found']
                ws.append(data_row)

    # Save the Excel file
    wb.save(output_excel)
    print(f"\nHypotheses summary saved to {output_excel}")

# Example usage:
# sem_analysis(model_desc, hypothesis_criteria)
